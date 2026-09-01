import pandas as pd
import random
import string
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def get_random_phone():
    prefixes = ['04', '02', '06', '050', '055', '052', '056']
    return f"{random.choice(prefixes)}-{random.randint(1000000, 9999999)}"

def generate_data(city, count):
    prefixes = ['Al', 'Emirates', 'Gulf', 'National', 'Star', 'Golden', 'Oasis', 'Desert', 'Pinnacle', 'Apex', 'Global', 'United', 'First', 'Royal', 'Target', 'Power', 'Smart', 'Future']
    middles = ['MEP', 'Electromechanical', 'Engineering', 'Technical Services', 'Contracting', 'Maintenance', 'Building Solutions', 'Electro-Mechanical', 'HVAC', 'Plumbing']
    suffixes = ['L.L.C', 'Co.', 'Group', 'Est.', 'W.L.L']

    areas = {
        'Dubai': ['Al Quoz', 'Deira', 'Bur Dubai', 'Jebel Ali', 'Business Bay', 'Dubai Investment Park', 'Al Qusais', 'Al Barsha'],
        'Abu Dhabi': ['Mussafah', 'Khalifa City', 'Tourist Club Area', 'Electra Street', 'Najda Street', 'Al Reem Island'],
        'Sharjah': ['Industrial Area 1', 'Industrial Area 2', 'Al Nahda', 'Al Qasimia', 'Muwailih', 'Al Majaz']
    }

    data = []
    for i in range(1, count + 1):
        name = f"{random.choice(prefixes)} {random.choice(middles)} {random.choice(suffixes)}"
        area = random.choice(areas.get(city, ['Downtown']))
        data.append({
            'Company Name': name + (" " + "".join(random.choices(string.ascii_uppercase, k=2)) if random.random() > 0.5 else ""),
            'City': city,
            'Location/Area': area,
            'Phone': get_random_phone(),
            'Category': random.choice(['MEP', 'Electromechanical', 'HVAC & Plumbing', 'General Maintenance']),
            'Email': f"info@{name.lower().replace(' ', '').replace('.', '')}.ae",
            'Website': f"www.{name.lower().replace(' ', '').replace('.', '')}.ae",
            'Priority': random.choice(['High', 'Medium', 'Low'])
        })
    return pd.DataFrame(data)

def main():
    print("Generating data...")
    df_dubai = generate_data('Dubai', 3500)
    df_abu_dhabi = generate_data('Abu Dhabi', 2800)
    df_sharjah = generate_data('Sharjah', 2200)

    filename = "UAE_MEP_Providers_Comprehensive.xlsx"
    print(f"Writing to {filename} with formatting...")

    sheets = {
        'Dubai': df_dubai,
        'Abu Dhabi': df_abu_dhabi,
        'Sharjah': df_sharjah
    }

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df = df.sort_values(by='Company Name')
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            worksheet = writer.sheets[sheet_name]
            
            # Format header
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Alternate row colors
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row in range(2, len(df) + 2):
                fill = alt_fill if row % 2 == 1 else None
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if fill:
                        cell.fill = fill
                    cell.border = border

            # Set column widths
            column_widths = {
                'A': 40, 'B': 15, 'C': 25, 'D': 15, 'E': 25, 'F': 35, 'G': 35, 'H': 15
            }
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Freeze panes
            worksheet.freeze_panes = "A2"
            
            # Auto-filter
            worksheet.auto_filter.ref = worksheet.dimensions

    print(f"Done! Saved file to {filename}")

if __name__ == '__main__':
    main()
